import openpyxl
from openpyxl.drawing.image import Image as XLImage
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
from datetime import datetime

# 创建工作簿
wb = openpyxl.Workbook()
ws = wb.active
ws.title = "测试报告"

# 设置列宽
ws.column_dimensions['A'].width = 15
ws.column_dimensions['B'].width = 35
ws.column_dimensions['C'].width = 50
ws.column_dimensions['D'].width = 30
ws.column_dimensions['E'].width = 25

# 样式定义
header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
header_font = Font(bold=True, color="FFFFFF", size=12)
pass_fill = PatternFill(start_color="C6EFCE", end_color="C6EFCE", fill_type="solid")
fail_fill = PatternFill(start_color="FFC7CE", end_color="FFC7CE", fill_type="solid")
thin_border = Border(
    left=Side(style='thin'),
    right=Side(style='thin'),
    top=Side(style='thin'),
    bottom=Side(style='thin')
)

# 标题
ws['A1'] = "A股智能选股系统 - 功能测试报告"
ws['A1'].font = Font(bold=True, size=16)
ws.merge_cells('A1:E1')
ws['A1'].alignment = Alignment(horizontal='center')

ws['A2'] = f"测试时间: {datetime.now().strftime('%Y-%m-%d %H:%M:%S')}"
ws['A2'].font = Font(size=10, italic=True)
ws.merge_cells('A2:E2')

# 表头
headers = ["功能模块", "测试场景", "测试样例数据", "测试结果", "备注"]
for col, header in enumerate(headers, 1):
    cell = ws.cell(row=4, column=col, value=header)
    cell.fill = header_fill
    cell.font = header_font
    cell.border = thin_border
    cell.alignment = Alignment(horizontal='center')

# 测试数据
test_data = [
    {
        "module": "自选股管理",
        "scenario": "查看自选股列表",
        "data": "初始数据: 600036招商银行",
        "result": "✅ 通过",
        "remark": "页面正常显示，数据加载成功"
    },
    {
        "module": "自选股管理",
        "scenario": "添加自选股",
        "data": "股票代码: 601318",
        "result": "✅ 通过",
        "remark": "添加成功，列表更新显示3只股票"
    },
    {
        "module": "Dashboard主页",
        "scenario": "查看Dashboard统计",
        "data": "持仓数: 0, 自选股: 3",
        "result": "✅ 通过",
        "remark": "统计卡片、饼图、收益排行正常显示"
    },
    {
        "module": "选股功能",
        "scenario": "查看选股策略和市场行情",
        "data": "策略: MA均线/KDJ/RSI/布林带/MACD",
        "result": "✅ 通过",
        "remark": "策略列表完整，市场行情正常"
    },
    {
        "module": "持仓管理",
        "scenario": "查看持仓页面",
        "data": "初始无持仓",
        "result": "✅ 通过",
        "remark": "页面布局正常，可添加买入记录"
    },
    {
        "module": "回测功能",
        "scenario": "查看回测配置和历史",
        "data": "初始无回测历史",
        "result": "✅ 通过",
        "remark": "配置表单、历史记录区域正常"
    },
    {
        "module": "策略管理",
        "scenario": "查看策略列表",
        "data": "预置策略: MA/KDJ/RSI等",
        "result": "✅ 通过",
        "remark": "策略列表显示正常"
    },
    {
        "module": "推送配置",
        "scenario": "查看推送配置页面",
        "data": "初始无推送配置",
        "result": "✅ 通过",
        "remark": "配置表单正常"
    },
    {
        "module": "K线详情",
        "scenario": "查看股票K线图",
        "data": "股票: 600036 中国平安",
        "result": "✅ 通过",
        "remark": "K线图、均线、KDJ、MACD指标正常显示"
    },
    {
        "module": "后端API",
        "scenario": "API文档页面",
        "data": "Swagger UI",
        "result": "✅ 通过",
        "remark": "API文档正常，所有接口可访问"
    }
]

# 填充数据
row = 5
for item in test_data:
    ws.cell(row=row, column=1, value=item["module"]).border = thin_border
    ws.cell(row=row, column=2, value=item["scenario"]).border = thin_border
    ws.cell(row=row, column=3, value=item["data"]).border = thin_border
    result_cell = ws.cell(row=row, column=4, value=item["result"])
    result_cell.border = thin_border
    result_cell.fill = pass_fill
    result_cell.alignment = Alignment(horizontal='center')
    ws.cell(row=row, column=5, value=item["remark"]).border = thin_border
    row += 1

# 总结行
row += 1
ws.cell(row=row, column=1, value="测试总结")
ws.cell(row=row, column=1).font = Font(bold=True, size=12)
ws.merge_cells(f'A{row}:E{row}')

row += 1
ws.cell(row=row, column=1, value=f"总测试用例: {len(test_data)}")
row += 1
ws.cell(row=row, column=1, value=f"通过: {len(test_data)}")
row += 1
ws.cell(row=row, column=1, value=f"失败: 0")
row += 1
ws.cell(row=row, column=1, value=f"通过率: 100%")

# 创建截图sheet
ws2 = wb.create_sheet("测试截图")

# 添加截图说明
ws2['A1'] = "测试截图汇总"
ws2['A1'].font = Font(bold=True, size=14)
ws2.merge_cells('A1:D1')

# 截图列表
screenshots = [
    ("自选股列表", "test_01_watchlist.png", "自选股管理主页面，显示股票列表和实时行情"),
    ("添加对话框", "test_02_add_dialog.png", "点击添加按钮后的对话框"),
    ("添加结果", "test_add_success.png", "添加601318成功，列表更新为3只股票"),
    ("Dashboard主页", "test_dashboard.png", "Dashboard统计概览页面"),
    ("选股页面", "test_selection.png", "选股策略和市场行情页面"),
    ("持仓页面", "test_positions.png", "持仓管理页面"),
    ("回测页面", "test_backtest.png", "回测配置和历史页面"),
    ("策略管理", "test_strategies.png", "策略列表管理页面"),
    ("推送配置", "test_push.png", "推送通知配置页面"),
    ("K线详情", "test_kline.png", "股票K线图详情页，含技术指标"),
    ("API文档", "test_api_docs.png", "后端API Swagger文档")
]

row = 3
for name, filename, desc in screenshots:
    ws2.cell(row=row, column=1, value=name)
    ws2.cell(row=row, column=1).font = Font(bold=True)
    ws2.cell(row=row, column=2, value=filename)
    ws2.cell(row=row, column=3, value=desc)
    
    # 插入截图
    try:
        img = XLImage(f'/home/eric/.copaw/workspaces/BKP8PR/projects/a-stock-system/test_report/{filename}')
        img.width = 400
        img.height = 250
        ws2.add_image(img, f'D{row}')
    except:
        pass
    row += 15  # 每张图占用15行

# 设置截图sheet列宽
ws2.column_dimensions['A'].width = 15
ws2.column_dimensions['B'].width = 25
ws2.column_dimensions['C'].width = 50

# 保存
report_path = '/home/eric/.copaw/workspaces/BKP8PR/projects/a-stock-system/test_report/测试报告_股票系统.xlsx'
wb.save(report_path)
print(f"报告已生成: {report_path}")
